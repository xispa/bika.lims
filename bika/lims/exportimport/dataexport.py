from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from Products.ATContentTypes.interfaces.interfaces import IATContentType
from openpyxl.style import NumberFormat, Border, Color, Font
from openpyxl.writer.styles import StyleWriter
from openpyxl.writer.excel import save_workbook
from Products.CMFCore.utils import getToolByName
from bika.lims import bikaMessageFactory as _
from bika.lims import logger
from bika.lims.utils import to_utf8
from bika.lims.utils import getFromString
from DateTime import DateTime
from Products.CMFPlone.utils import safe_unicode
import tempfile

CATALOG = 'catalog'
EXCLUDE = 'exclude'
INCLUDE = 'include'
SHEET_NAME = 'sheet-name'
QUERY = 'query'

DEFAULT_FIELDS_TO_OMIT = ['id', 'constrainTypesMode', 'allowDiscussion',
                          'excludeFromNav', 'nextPreviousEnabled', 'location',
                          'language', 'effectiveDate', 'expirationDate',
                          'rights', 'creation_date', 'modification_date',
                          'ObjectWorkflowStates']
DEFAULT_PORTAL_TYPES = {

    'Client':
        {CATALOG: 'portal_catalog',
         SHEET_NAME: 'Clients',
         EXCLUDE: ['title', 'description',]},

    'Contact':
        {CATALOG: 'portal_catalog',
         SHEET_NAME: 'Client Contacts',
         EXCLUDE: ['title', 'description', 'Fullname'],
         INCLUDE: ['aq_parent.Title:Client_title']},

    'AnalysisCategory':
        {CATALOG: 'bika_setup_catalog',
         SHEET_NAME: 'Analysis Categories',
         EXCLUDE: ['SortKey', 'DepartmentTitle'],
         INCLUDE: []},

    'AnalysisService':
        {CATALOG: 'bika_setup_catalog',
         SHEET_NAME: 'Analysis Services',
         EXCLUDE: ['SortKey', 'Category', 'Department'],
         INCLUDE: ['getCategory.Title:AnalysisCategory_title',
                   'getDepartment.Title:Department_title']}
}

class SetupDataExporter:

    def __init__(self, context):
        self.workbook = None
        self.ws_cols = {}
        self.context = context


    def export(self, portal_types=None, out_file_path=None):
        if not portal_types:
            portal_types = DEFAULT_PORTAL_TYPES

        for portal_type, options in portal_types.iteritems():
            catalog = options.get(CATALOG, 'portal_catalog')
            exclude = options.get(EXCLUDE, [])
            include = options.get(INCLUDE, [])
            query = options.get(QUERY, None)

            self._to_worksheet(portal_type, catalog, query, exclude, include)

        if self.workbook:
            # Apply styles
            w = StyleWriter(self.workbook)
            w._write_fonts()
            w._write_fills()

            # Write file
            outfile = tempfile.mktemp(suffix=".xlsx")
            if out_file_path:
                outfile = out_file_path
            dest_filename = outfile
            logger.info("Saving export data file: %s" % dest_filename)
            save_workbook(self.workbook, dest_filename)


    def _to_worksheet(self, portal_type, catalog='portal_catalog', query=None,
                     exclude_fields=None, include_fields=None):
        if not exclude_fields:
            exclude_fields=[]
        if not include_fields:
            include_fields=[]
        ws = self._get_worksheet(portal_type)
        is_empty = True
        tool = getToolByName(self.context, catalog)
        brains = tool(portal_type=portal_type, sort_on='sortable_title',
                      sort_order='ascending')
        for brain in brains:
            obj = brain.getObject()
            review_state = ''
            if hasattr(brain, 'review_state'):
                review_state = brain.review_state
            if is_empty:
                # This is an empty worksheet, fill the first and second rows
                self._fill_headers(ws, obj, exclude_fields, include_fields)
                is_empty = False
            self._append_row(ws, obj, review_state)

        # Adjust the width of the columns
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if type(cell.value) is not unicode:
                        continue
                    width_cell = len(cell.value) * 0.8
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), width_cell))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value


    def _get_worksheet(self, portal_type):
        if not self.workbook:
            self.workbook = Workbook()
        sheet_name = DEFAULT_PORTAL_TYPES.get(portal_type, {})
        sheet_name = sheet_name.get(SHEET_NAME, portal_type)
        ws = self.workbook.get_sheet_by_name(sheet_name)
        if not ws:
            ws = self.workbook.create_sheet(title=sheet_name)
        return ws


    def _fill_headers(self, ws, obj, exclude_fields=None, include_fields=None):
        if not exclude_fields:
            exclude_fields = []
        if not include_fields:
            include_fields = []
        schema = obj.Schema()
        cols = []
        heads = []
        wcols = []
        for field in include_fields:
            tokens = field.split(':')
            if len(tokens) > 1:
                cols.append(tokens[1])
                heads.append(_(tokens[1]))
                wcols.append(tokens[0])
            else:
                cols.append(field)
                heads.append(_(field))
                wcols.append(field)

        for field in schema.fields():
            field_name = field.getName()
            if field_name in exclude_fields or \
                            field_name in DEFAULT_FIELDS_TO_OMIT:
                continue
            if field.type == 'lines':
                # Dismiss tuples for now
                continue
            if field.type == 'reference' and field.multiValued:
                # Dismiss multivalued references for now
                continue
            if field.type == 'address':
                # Need to create multiple columns for this
                for subfield in field.subfields:
                    subfield_name = '%s_%s' % (field_name, subfield)
                    cols.append(subfield_name)
                    wcols.append(subfield_name)
                    heads.append(_(subfield_name))
                continue

            cols.append(field_name)
            heads.append(_(field_name))
            wcols.append(field_name)

        cols.append('review_state')
        heads.append(_('review_state'))
        wcols.append('review_state')
        ws.append(cols)
        ws.append([ws.title])
        ws.cell('A2').style.font.size = 12
        ws.cell('A2').style.font.name='Arial'
        ws.cell('A2').style.font.bold = True
        ws.cell('A2').style.alignment.shrink_to_fit = True
        ws.append(heads)
        for idx in range(0, len(heads)):
            letter = get_column_letter(idx + 1)
            col_coord = '%s3' % letter
            ws.cell(col_coord).style.font.bold = True
            ws.cell(col_coord).style.font.name='Arial'
            ws.cell(col_coord).style.alignment.shrink_to_fit = True
            col_coord = '%s1' % letter
            ws.cell(col_coord).style.font.size=8
            ws.cell(col_coord).style.font.name='Arial'
            ws.cell(col_coord).style.alignment.shrink_to_fit = True
        self.ws_cols[ws.title] = wcols


    def _append_row(self, ws, obj, review_state):
        columns = self.ws_cols[ws.title]
        row = []
        for col in columns:
            if col == 'review_state':
                row.append(review_state)
                continue
            field = obj.getField(col, '')
            if not field:
                # Is an include_field?
                value = getFromString(obj, col)
                if not value:
                    # Is a multicolumn field (e.g. address)?
                    tokens = col.split('_')
                    if len(tokens) > 1:
                        field = obj.getField(tokens[0], '')
                        value = field.get(obj) if field else {}
                        value = value.get(tokens[1], '')
                        value = to_utf8(value)
                        row.append(value)
                        continue
            else:
                value = field.get(obj)

            value = value if value else ''
            # Check field type?
            if field and field.type == 'boolean':
                value = '1' if value else '0'
            elif type(value) is DateTime:
                # Convert to YYYY-MM-dd hh:mm:ss
                value = value.asdatetime()
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif field and field.type == 'reference' and not field.multiValued:
                value = to_utf8(value.Title())
            elif type(value) is IATContentType:
                value = to_utf8(value.Title())
            elif not value:
                value = ''
            else:
                value = to_utf8(str(value))
            row.append(value)
        ws.append(row)
        row_idx = len(ws.row_dimensions)
        for idx in range(0, len(row)):
            letter = get_column_letter(idx + 1)
            col_coord = '%s%s' % (letter, row_idx)
            cell = ws.cell(col_coord)
            cell.style.font.size=10
            cell.style.font.name='Arial'
            cell.style.alignment.shrink_to_fit = True

